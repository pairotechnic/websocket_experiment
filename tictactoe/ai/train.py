"""
Self-play Q-learning trainer.
Run from the /tictactoe directory:
    python -m ai.train
"""

# Standard Library Imports
from datetime import datetime
import math
from pathlib import Path
import uuid

# Third-Party Library Imports
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Local Application Imports
from ai.q_agent import QAgent
from game import Game

SAVE_PATH = (
    Path(__file__).parent
    / "models"
    / f"q_table_self_{datetime.now().strftime('%Y_%m_%d_%H_%M_%S')}.pkl"
)

EPISODES = 200_000
LOG_EVERY = 10_000
LOG_PATH = Path(__file__).parent / "models" / "_training_log.xlsx"
EVALUATION_EPISODES = 10_000

# Rewards
REWARD_WIN = 1.0
REWARD_LOSE = -1.0
REWARD_DRAW = 0.5      # positive: prefer drawing to losing
REWARD_STEP = 0.0      # no reward for non-terminal moves


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

def log_run(agent: QAgent, episodes: int, wins: dict, q_table_size: int):
    cols = [
        "model filename", "alpha", "gamma", "epsilon start", "epsilon min",
        "epsilon decay", "epsilon min episode", "episodes",
        "x wins", "o wins", "draws", "q-table size",
    ]

    if LOG_PATH.exists():
        wb = openpyxl.load_workbook(LOG_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Training Log"
        header_fill = PatternFill("solid", start_color="1F4E79")
        for c, col in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=c, value=col)
            cell.font = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        col_widths = [36, 8, 8, 14, 12, 14, 20, 10, 10, 10, 10, 14]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    if agent.epsilon_decay < 1.0:
        ep_min = math.ceil(
            math.log(agent.epsilon_min / 1.0) / math.log(agent.epsilon_decay)
        )
        ep_min = min(ep_min, episodes)
    else:
        ep_min = episodes

    row = [
        SAVE_PATH.name,
        agent.alpha,
        agent.gamma,
        1.0,
        agent.epsilon_min,
        agent.epsilon_decay,
        ep_min,
        episodes,
        wins["X"],
        wins["O"],
        wins["Draw"],
        q_table_size,
    ]

    next_row = ws.max_row + 1
    fill = PatternFill("solid", start_color="D6E4F0") if next_row % 2 == 0 else None
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=next_row, column=c, value=val)
        cell.font = Font(name="Arial")
        cell.alignment = Alignment(horizontal="center")
        if fill:
            cell.fill = fill

    wb.save(LOG_PATH)
    print(f"Logged run -> {LOG_PATH}")


# ---------------------------------------------------------------------------
# Episode runner
# ---------------------------------------------------------------------------

def run_episode(agent: QAgent, is_training: bool = False) -> str | None:
    """
    Play one full game. A single shared agent plays both sides; the board is
    canonicalized per-turn so the Q-table is always from the moving player's
    perspective.

    Returns "X", "O", or None (draw).
    """
    game = Game(game_id=str(uuid.uuid4())[:8])
    symbols = ["X", "O"]

    # Clear any state left over from the previous episode
    agent.reset_episode()

    # We need to remember the last acting symbol so we can issue a deferred
    # update to the *previous* mover after the *current* mover's action lands.
    prev_symbol: str | None = None

    while not game.game_over:
        idx = game.current_turn          # 0 = X, 1 = O
        my_symbol = symbols[idx]

        action = agent.act(game.board, my_symbol)
        game.make_move(idx, action)

        if not is_training:
            continue

        if game.game_over:
            if game.winner == my_symbol:
                # Current mover won
                agent.learn(game.board, my_symbol, REWARD_WIN, done=True)
                # Previous mover lost — give them a deferred losing update
                if prev_symbol is not None:
                    agent.learn(game.board, prev_symbol, REWARD_LOSE, done=True)
            else:
                # Draw (winner is None; a loss mid-game can't happen here
                # because make_move only sets game_over on win or full board)
                agent.learn(game.board, my_symbol, REWARD_DRAW, done=True)
                if prev_symbol is not None:
                    agent.learn(game.board, prev_symbol, REWARD_DRAW, done=True)
        else:
            # Non-terminal: give the *previous* mover their deferred update now
            # that we can see what the opponent did in response.
            if prev_symbol is not None:
                agent.learn(game.board, prev_symbol, REWARD_STEP, done=False)

            # The current mover's update will come next turn (deferred).
            # We do need to call act() again next iteration, which will
            # overwrite last_state/last_action — so snapshot them first
            # by doing nothing: learn() is intentionally called *after*
            # the next act(), using the saved last_state/last_action.

        prev_symbol = my_symbol

    return game.winner


# ---------------------------------------------------------------------------
# Training loop
# ---------------------------------------------------------------------------

def train(agent: QAgent):
    wins = {"X": 0, "O": 0, "Draw": 0}
    total_wins = {"X": 0, "O": 0, "Draw": 0}

    for ep in range(1, EPISODES + 1):
        winner = run_episode(agent, is_training=True)
        key = winner if winner else "Draw"
        wins[key] += 1
        total_wins[key] += 1

        agent.decay_epsilon()

        if ep % LOG_EVERY == 0:
            total = sum(wins.values())
            print(
                f"Episode {ep:>7} | "
                f"ε={agent.epsilon:.4f} | "
                f"X wins: {wins['X']/total:.1%}  "
                f"O wins: {wins['O']/total:.1%}  "
                f"Draws: {wins['Draw']/total:.1%} | "
                f"Q-table size: {len(agent.q_table)}"
            )
            wins = {"X": 0, "O": 0, "Draw": 0}

    agent.save(SAVE_PATH)
    log_run(agent, EPISODES, total_wins, len(agent.q_table))
    print("Training complete.")

    _diagnostics(agent)


# ---------------------------------------------------------------------------
# Evaluation
# ---------------------------------------------------------------------------

def evaluate(agent: QAgent):
    """Greedy self-play evaluation — no exploration."""
    saved_epsilon = agent.epsilon
    agent.epsilon = 0.0

    wins = {"X": 0, "O": 0, "Draw": 0}
    for _ in range(1, EVALUATION_EPISODES + 1):
        winner = run_episode(agent, is_training=False)
        key = winner if winner else "Draw"
        wins[key] += 1

    total = sum(wins.values())
    print(
        f"Evaluation ({EVALUATION_EPISODES} eps) | "
        f"X wins: {wins['X']/total:.1%}  "
        f"O wins: {wins['O']/total:.1%}  "
        f"Draws: {wins['Draw']/total:.1%} | "
        f"Q-table size: {len(agent.q_table)}"
    )

    agent.epsilon = saved_epsilon


# ---------------------------------------------------------------------------
# Diagnostics
# ---------------------------------------------------------------------------

def _diagnostics(agent: QAgent):
    # Diagnostic 1 — Q-values for the canonical empty board (agent's first move)
    print("\nDiagnostic 1 — Q-values from empty board (canonical, agent moves first)")
    empty = tuple([""] * 9)
    for action in range(9):
        print(f"  cell {action}: {agent.q_table.get((empty, action), 'missing')}")

    # Diagnostic 2 — sparsity
    print("\nDiagnostic 2 — Q-table sparsity")
    count_nonzero = sum(1 for v in agent.q_table.values() if v != 0)
    print(f"  Total entries    : {len(agent.q_table)}")
    print(f"  Non-zero entries : {count_nonzero}")

    # Diagnostic 3 — value range
    print("\nDiagnostic 3 — Q-value range")
    max_q = max(agent.q_table.values())
    min_q = min(agent.q_table.values())
    print(f"  max_q = {max_q:.4f}")
    print(f"  min_q = {min_q:.4f}")

    # Diagnostic 4 — top 10 entries
    print("\nDiagnostic 4 — Top 10 Q-values")
    largest = sorted(agent.q_table.items(), key=lambda kv: kv[1], reverse=True)[:10]
    for (state, action), value in largest:
        moves_played = sum(1 for cell in state if cell != "")
        print(f"  moves={moves_played} action={action} q={value:.3f} state={state}")

    # Diagnostic 5 — canonical second-move responses
    # (opponent played corner/centre/edge; what does the agent prefer?)
    print("\nDiagnostic 5 — Best canonical response to opponent's first move")
    opp_first_moves = {
        "opponent corner (0)": tuple(["O", "", "", "", "", "", "", "", ""]),
        "opponent centre (4)": tuple(["", "", "", "", "O", "", "", "", ""]),
        "opponent edge   (1)": tuple(["", "O", "", "", "", "", "", "", ""]),
    }
    for label, state in opp_first_moves.items():
        legal = [i for i, c in enumerate(state) if c == ""]
        best = max(legal, key=lambda a: agent.q_table.get((state, a), 0.0))
        best_val = agent.q_table.get((state, best), 0.0)
        print(f"  {label}  -> best cell={best}  q={best_val:.4f}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    # A single agent plays both sides; the canonical board view means its
    # Q-table is symbol-agnostic and transfers perfectly to inference time
    # regardless of whether the human picks X or O.
    agent = QAgent()
    train(agent)
    evaluate(agent)


if __name__ == "__main__":
    main()